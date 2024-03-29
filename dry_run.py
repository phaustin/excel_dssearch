import context
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import copy
from dataclasses import dataclass


@dataclass
class RevInfo:
    """
    initial_dict maps reviewer names to initials
    range_dict maps applicant columns to sheet range names
    item_dict maps range names to ranges
    rev_dict maps reviewer column and filename to their values
    """

    initial_dict: dict
    range_dict: dict
    item_dict: dict
    rev_dict: dict


def assign_reviewer(row, rev_info):
    """
    read a candidate's row  which has
    reviewer number 1 and 2 indicated under the names column and return
    a dictionary of the form {'rev_1': 'rp', 'rev_2': 'sw'}
    """
    reviewer_dict = {}
    #
    # intial_dict.keys() are [1,2]
    #
    for reviewer in rev_info.initial_dict.keys():
        if row[reviewer] > 0:
            reviewer_val = int(row[reviewer])
            reviewer_head = rev_info.rev_dict[reviewer_val]["rev"]
            reviewer_dict[reviewer_head] = rev_info.initial_dict[reviewer]
    return reviewer_dict


def make_filename(row):
    """
    create 2 filenames from the applicant and the reviewer of the form
    {'rev1_file': 'aa/candname-revnum_revinitials.xlsx',
               'rev2_file': 'candname-revnum_revinitials.xlsx'}
    """
    the_name = row["Applicant Name"]
    the_name = the_name.replace(" ", "_")
    the_name = the_name.replace(",", "_")
    rev2_name = f'{row["rev_2"]}/{the_name}-2_{row["rev_2"]}.xlsx'
    rev1_name = f'{row["rev_1"]}/{the_name}-1_{row["rev_1"]}.xlsx'
    reviewer_dict = dict(rev1_file=rev1_name, rev2_file=rev2_name)
    print(reviewer_dict)
    return reviewer_dict


def make_xlfile(filename, base_dir):
    """
    create a full Path object with the relative filename
    and the base path
    """
    the_file = Path(filename)
    rev_dir = the_file.parent
    out_dir = base_dir / rev_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    outfile = out_dir / the_file.name
    return outfile


def fill_blanks(sheet, row, index, reviewer_num, rev_info):
    """
    fill an individual candidate spreadsheet with
    cell information
    """
    rev_col = rev_info.rev_dict[reviewer_num]["rev"]
    initials = row[rev_col]
    cellnum = rev_info.item_dict["initials"]
    sheet[cellnum] = initials
    cellnum = rev_info.item_dict["id"]
    sheet[cellnum] = index
    for item in ["name", "school", "year", "level"]:
        col_name = rev_info.range_dict[item]
        value = row[col_name]
        cellnum = rev_info.item_dict[item]
        sheet[cellnum] = value
    return sheet


if __name__ == "__main__":
    #
    # find the template file and extract the named ranges
    #
    excel_orig = context.dsci_search / Path("draft_evaluation_form.xlsx")
    print(f"template is {str(excel_orig)}")

    #
    # range_dict maps applicant column names to candidate sheet range names
    #
    range_names = ["name", "school", "year", "level"]
    colnames = [
        "Applicant Name",
        "School (PhD)",
        "Year (PhD)",
        "Highest Education Level",
    ]
    range_dict = dict(zip(range_names, colnames))
    #
    # item_dict stores the named range values for the individual sheet cells
    #
    wb_copy = load_workbook(filename=str(excel_orig))
    range_names.extend(["initials", "id"])
    item_dict = {}
    for item in range_names:
        the_range = wb_copy.defined_names[item]
        row_col = list(the_range.destinations)[0][1]
        item_dict[item] = row_col

    rev_dict = {
        1: {"rev": "rev_1", "filename": "rev1_file"},
        2: {"rev": "rev_2", "filename": "rev2_file"},
    }

    names = ["Pawlowicz", "Ameli", "Austin", "Haber", "Waterman"]
    initials = ["rp", "aa", "pa", "eh", "sw"]
    initial_dict = dict(zip(names, initials))
    rev_info = RevInfo(initial_dict, range_dict, item_dict, rev_dict)

    #
    # get the list of candidates from the applicant list spreadsheewt
    #
    candidate_file = context.dsci_search / Path("EDS - applicant list.xlsx")
    print(f'reading "{str(candidate_file)}"')
    df_candidates = pd.read_excel(str(candidate_file), skiprows=[0])
    columns = [
        "Applicant Name",
        "School (PhD)",
        "Year (PhD)",
        "Highest Education Level",
        "Pawlowicz",
        "Ameli",
        "Austin",
        "Haber",
        "Waterman",
    ]
    #
    # subset the columns
    #
    df_candidates.fillna(0, inplace=True)
    df_candidates = pd.DataFrame(df_candidates[columns], copy=True)
    print(df_candidates.index.values)

    #
    # add columns for reviewer 1 and 2 initials and output file names
    #
    out = df_candidates.apply(assign_reviewer, args=(rev_info,), axis=1)
    df_revs = pd.DataFrame.from_records(out)
    df_candidates[["rev_2", "rev_1"]] = df_revs[["rev_2", "rev_1"]]
    out = df_candidates.apply(make_filename, axis=1)
    df_revs = pd.DataFrame.from_records(out)
    df_candidates[["rev2_file", "rev1_file"]] = df_revs[["rev2_file", "rev1_file"]]

    base_dir = Path() / "sheets"
    all_rows = list(df_candidates.iterrows())
    for index, row in all_rows:
        for rev_key in [1, 2]:
            wb_copy = load_workbook(filename=str(excel_orig))
            the_sheet = wb_copy["main"]
            filled_sheet = fill_blanks(the_sheet, row, index, rev_key, rev_info)
            name_col = rev_dict[rev_key]["filename"]
            filename = row[name_col]
            outfile = make_xlfile(filename, base_dir)
            wb_copy.save(outfile)
